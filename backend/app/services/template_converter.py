"""Convert uploaded template files (Excel, CSV, DOCX) to HTML."""

import base64
from io import BytesIO
from html import escape
from xml.etree.ElementTree import fromstring

import openpyxl
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter, range_boundaries


# ---------------------------------------------------------------------------
# Theme color resolution
# ---------------------------------------------------------------------------

# XML tag order in <a:clrScheme>: dk1, lt1, dk2, lt2, accent1..6, hlink, folHlink
# openpyxl swaps indices 0↔1 and 2↔3 relative to the XML order.
_THEME_XML_ORDER = [
    "dk1", "lt1", "dk2", "lt2",
    "accent1", "accent2", "accent3", "accent4",
    "accent5", "accent6", "hlink", "folHlink",
]
_OPENPYXL_TO_XML = {0: 1, 1: 0, 2: 3, 3: 2}  # swap pairs; rest identity


def _parse_theme_colors(wb) -> list[str]:
    """Extract the 12 theme colors from the workbook's theme XML.

    Returns a list of 12 hex strings (e.g. "FFFFFF") indexed by openpyxl's
    theme index convention (0=lt1, 1=dk1, 2=lt2, 3=dk2, 4..9=accent1..6,
    10=hlink, 11=folHlink).
    """
    theme_data = getattr(wb, "loaded_theme", None)
    if theme_data is None:
        return []
    try:
        root = fromstring(theme_data)
    except Exception:
        return []
    ns = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    scheme = root.find(".//a:clrScheme", ns)
    if scheme is None:
        return []

    # Build list in XML document order
    xml_colors: list[str] = []
    for child in scheme:
        tag = child.tag.split("}")[-1]
        for sub in child:
            val = sub.get("val") or sub.get("lastClr")
            if val:
                xml_colors.append(val.upper())
                break
        else:
            xml_colors.append("000000")

    # Reorder so index matches openpyxl's convention (swap 0↔1, 2↔3)
    reordered: list[str] = list(xml_colors)
    for opx_idx, xml_idx in _OPENPYXL_TO_XML.items():
        if xml_idx < len(xml_colors) and opx_idx < len(reordered):
            reordered[opx_idx] = xml_colors[xml_idx]
    return reordered


def _apply_tint(hex_rgb: str, tint: float) -> str:
    """Apply an Excel tint value to a hex RGB string.

    Positive tint → lighter (toward white).
    Negative tint → darker (toward black).
    """
    if tint == 0.0:
        return hex_rgb
    r = int(hex_rgb[0:2], 16)
    g = int(hex_rgb[2:4], 16)
    b = int(hex_rgb[4:6], 16)

    def _tint_channel(c: int) -> int:
        if tint > 0:
            return int(c + (255 - c) * tint)
        return int(c * (1 + tint))

    r, g, b = _tint_channel(r), _tint_channel(g), _tint_channel(b)
    return f"{min(max(r,0),255):02X}{min(max(g,0),255):02X}{min(max(b,0),255):02X}"


def _resolve_color(color, theme_colors: list[str]) -> str | None:
    """Resolve an openpyxl Color object to a CSS hex string, or None."""
    if color is None:
        return None

    color_type = getattr(color, "type", None)

    # --- theme color ---
    if color_type == "theme":
        idx = color.theme
        if isinstance(idx, int) and 0 <= idx < len(theme_colors):
            base = theme_colors[idx].upper()
            # Resolve system color names to hex
            system_map = {"WINDOW": "FFFFFF", "WINDOWTEXT": "000000"}
            if base in system_map:
                base = system_map[base]
            if not all(c in "0123456789ABCDEF" for c in base) or len(base) != 6:
                return None
            tint = getattr(color, "tint", 0.0) or 0.0
            return f"#{_apply_tint(base, tint)}"
        return None

    # --- indexed color ---
    if color_type == "indexed":
        idx = getattr(color, "index", None)
        if idx is None:
            idx = getattr(color, "indexed", None)
        if isinstance(idx, int) and 0 <= idx < len(COLOR_INDEX):
            raw = COLOR_INDEX[idx]
            if raw == "00000000":
                return None
            if len(raw) == 8:
                raw = raw[2:]
            return f"#{raw}"
        return None

    # --- explicit rgb ---
    if color_type == "rgb":
        try:
            rgb = color.rgb
        except Exception:
            return None
        if rgb is None:
            return None
        rgb = str(rgb)
        # Guard against openpyxl error strings leaking through
        if not all(c in "0123456789ABCDEFabcdef" for c in rgb):
            return None
        if rgb in ("00000000", "0"):
            return None
        if len(rgb) == 8:
            rgb = rgb[2:]
        return f"#{rgb}"

    # --- fallback: try reading .rgb directly ---
    try:
        rgb = color.rgb
        if rgb is None:
            return None
        rgb = str(rgb)
        if not all(c in "0123456789ABCDEFabcdef" for c in rgb):
            return None
        if rgb in ("00000000", "0"):
            return None
        if len(rgb) == 8:
            rgb = rgb[2:]
        return f"#{rgb}"
    except Exception:
        return None


def _border_css(side, theme_colors: list[str]) -> str:
    if side is None or side.style is None:
        return "none"
    weight = {"thin": "1px", "medium": "2px", "thick": "3px"}.get(side.style, "1px")
    color = _resolve_color(side.color, theme_colors) if side.color else "#000"
    return f"{weight} solid {color or '#000'}"


def _cell_style(cell, theme_colors: list[str], scale: float = 1.0) -> str:
    parts: list[str] = []
    if cell.font:
        if cell.font.bold:
            parts.append("font-weight:bold")
        if cell.font.size:
            parts.append(f"font-size:{cell.font.size * scale:.1f}pt")
        fc = _resolve_color(cell.font.color, theme_colors) if cell.font.color else None
        if fc:
            parts.append(f"color:{fc}")
    if cell.fill and cell.fill.fgColor:
        bg = _resolve_color(cell.fill.fgColor, theme_colors)
        if bg:
            parts.append(f"background-color:{bg}")
    if cell.alignment:
        if cell.alignment.horizontal:
            parts.append(f"text-align:{cell.alignment.horizontal}")
        if cell.alignment.vertical:
            va = cell.alignment.vertical
            if va == "center":
                va = "middle"
            parts.append(f"vertical-align:{va}")
    if cell.border:
        b = cell.border
        parts.append(f"border-top:{_border_css(b.top, theme_colors)}")
        parts.append(f"border-right:{_border_css(b.right, theme_colors)}")
        parts.append(f"border-bottom:{_border_css(b.bottom, theme_colors)}")
        parts.append(f"border-left:{_border_css(b.left, theme_colors)}")
    return ";".join(parts)


def _anchor_display_size(ws, anchor) -> tuple[int, int]:
    """Calculate display width and height in pixels from a TwoCellAnchor.

    Anchor row/col values are 0-based.  Offsets are in EMUs (914400 per inch).
    """
    EMU_PER_INCH = 914400
    PX_PER_INCH = 96
    PT_PER_INCH = 72

    f = anchor._from
    t = anchor.to

    # --- height: sum row heights between from.row and to.row, adjust offsets ---
    total_h_emu = 0
    for r in range(f.row, t.row):
        # row dimensions use 1-based indexing
        rd = ws.row_dimensions.get(r + 1)
        h_pt = (rd.height if rd and rd.height else 15.0)
        total_h_emu += int(h_pt * EMU_PER_INCH / PT_PER_INCH)
    total_h_emu = total_h_emu - f.rowOff + t.rowOff

    # --- width: sum column widths between from.col and to.col, adjust offsets ---
    total_w_emu = 0
    for c in range(f.col, t.col):
        # column dimensions use letter keys; anchor cols are 0-based
        letter = get_column_letter(c + 1)
        dim = ws.column_dimensions.get(letter)
        w_chars = (dim.width if dim and dim.width else 8.43)
        # Excel character-width to pixels: px ≈ (chars * 7 + 5)
        # Then pixels to EMU
        w_px = int(w_chars * 7 + 5)
        total_w_emu += int(w_px * EMU_PER_INCH / PX_PER_INCH)
    total_w_emu = total_w_emu - f.colOff + t.colOff

    w_px = max(1, int(total_w_emu * PX_PER_INCH / EMU_PER_INCH))
    h_px = max(1, int(total_h_emu * PX_PER_INCH / EMU_PER_INCH))
    return w_px, h_px


def _extract_images(ws) -> dict[tuple[int, int], str]:
    """Build a map of (row, col) -> base64 <img> tag for embedded images."""
    images: dict[tuple[int, int], str] = {}
    for img in ws._images:
        anchor = img.anchor
        if not hasattr(anchor, "_from"):
            continue
        row = anchor._from.row
        col = anchor._from.col
        try:
            raw = img._data()
        except Exception:
            continue
        fmt = getattr(img, "format", "png") or "png"
        b64 = base64.b64encode(raw).decode("ascii")

        # Calculate display size from TwoCellAnchor cell positions
        if hasattr(anchor, "to"):
            w_px, h_px = _anchor_display_size(ws, anchor)
        else:
            # Fallback for OneCellAnchor or missing 'to': use raw dims
            w_px = int(img.width * 96 / 914400) if img.width > 1000 else img.width
            h_px = int(img.height * 96 / 914400) if img.height > 1000 else img.height

        tag = (
            f'<img src="data:image/{fmt};base64,{b64}" '
            f'style="width:{w_px}px;height:{h_px}px;object-fit:contain">'
        )
        images[(row, col)] = tag
    return images


def _get_print_bounds(ws) -> tuple[int, int, int, int]:
    """Return (min_col, min_row, max_col, max_row) from print area, or sheet dims."""
    pa = ws.print_area
    if pa:
        ref = pa.split("!")[-1].replace("$", "")
        return range_boundaries(ref)
    return 1, 1, ws.max_column, ws.max_row


def _row_height_pt(ws, row: int) -> float:
    if row in ws.row_dimensions and ws.row_dimensions[row].height:
        return ws.row_dimensions[row].height
    return 15.0


def _col_width_chars(ws, col: int) -> float:
    letter = get_column_letter(col)
    dim = ws.column_dimensions.get(letter)
    if dim and dim.width:
        return dim.width
    return 8.43


def excel_to_html(data: bytes, return_extractions: bool = False) -> str | tuple[str, list]:
    """Convert an Excel workbook to HTML.

    When *return_extractions* is True, returns a tuple of (html, extractions)
    where extractions is a list of FormulaExtraction dicts from formula_extractor.
    """
    from app.services.formula_extractor import extract_formulas

    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb.active
    theme_colors = _parse_theme_colors(wb)

    # Pre-compute formula extractions keyed by cell reference
    raw_extractions = extract_formulas(wb)
    extraction_map: dict[str, object] = {
        ext.cell_ref: ext for ext in raw_extractions
    }

    min_col, min_row, max_col, max_row = _get_print_bounds(ws)

    # Trim trailing empty rows
    while max_row > min_row:
        has_content = False
        for c in range(min_col, max_col + 1):
            if ws.cell(row=max_row, column=c).value is not None:
                has_content = True
                break
        if has_content:
            break
        max_row -= 1

    # Read margins (inches) — default to 0.7 left/right, 0.75 top/bottom
    margin_top = ws.page_margins.top if ws.page_margins.top else 0.75
    margin_bottom = ws.page_margins.bottom if ws.page_margins.bottom else 0.75
    margin_left = ws.page_margins.left if ws.page_margins.left else 0.7
    margin_right = ws.page_margins.right if ws.page_margins.right else 0.7

    # Read print scale (Excel default is 100)
    scale_pct = ws.page_setup.scale if ws.page_setup.scale else 100
    scale = scale_pct / 100.0

    # Build merge map within print area
    merge_map: dict[str, tuple[int, int]] = {}
    skip_cells: set[str] = set()
    for mr in ws.merged_cells.ranges:
        if mr.min_col < min_col or mr.max_col > max_col:
            continue
        if mr.min_row < min_row or mr.max_row > max_row:
            continue
        rowspan = mr.max_row - mr.min_row + 1
        colspan = mr.max_col - mr.min_col + 1
        anchor = f"{mr.min_row},{mr.min_col}"
        merge_map[anchor] = (rowspan, colspan)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if r != mr.min_row or c != mr.min_col:
                    skip_cells.add(f"{r},{c}")

    # Extract embedded images
    image_map = _extract_images(ws)

    # Column widths as percentages of total
    raw_widths = [_col_width_chars(ws, c) for c in range(min_col, max_col + 1)]
    total_w = sum(raw_widths)
    pct_widths = [w / total_w * 100 for w in raw_widths]

    # Determine which leading rows are image-only (no text in any cell).
    # These get extracted into a flex div above the table so their colspan
    # images don't distort data column widths.
    header_end = min_row
    for row in range(min_row, max_row + 1):
        has_text = False
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                has_text = True
                break
        if has_text:
            break
        header_end = row + 1

    # Collect images that fall within the extracted header rows
    header_images: list[str] = []
    if header_end > min_row:
        for img_key in sorted(image_map):
            img_row_0based, _ = img_key
            img_row = img_row_0based + 1
            if min_row <= img_row < header_end:
                header_images.append(image_map[img_key])

    # Compute total height of extracted header rows for the flex container
    header_height_pt = sum(
        _row_height_pt(ws, r) for r in range(min_row, header_end)
    )

    # Page CSS
    page_css = (
        f"@page {{"
        f" size: letter portrait;"
        f" margin: {margin_top}in {margin_right}in {margin_bottom}in {margin_left}in;"
        f"}}"
    )

    base_font_pt = 11.0 * scale
    lines = [
        f"<style>{page_css}",
        f"body {{ margin: 0; padding: 0; font-family: Calibri, Arial, sans-serif; font-size: {base_font_pt:.1f}pt; }}",
        ".header-images { display: flex; align-items: center; justify-content: space-between; }",
        ".header-images img { max-height: 100%; }",
        f"table {{ width: 100%; border-collapse: collapse; font-family: Calibri, Arial, sans-serif; font-size: {base_font_pt:.1f}pt; }}",
        "td { padding: 2px 4px; }",
        ".formula-unrecognized { color: #999; font-style: italic; }",
        "</style>",
    ]

    if header_images:
        lines.append(
            f'<div class="header-images" style="height:{header_height_pt * scale:.1f}pt">'
        )
        for img_tag in header_images:
            lines.append(f"  {img_tag}")
        lines.append("</div>")

    lines.append("<table>")
    lines.append("<colgroup>")
    for pct in pct_widths:
        lines.append(f'  <col style="width:{pct:.1f}%">')
    lines.append("</colgroup>")
    lines.append("<tbody>")

    table_start = header_end if header_end > min_row else min_row

    for row in range(table_start, max_row + 1):
        rh = _row_height_pt(ws, row) * scale
        lines.append(f'<tr style="height:{rh:.1f}pt">')

        for col in range(min_col, max_col + 1):
            key = f"{row},{col}"
            if key in skip_cells:
                continue

            cell = ws.cell(row=row, column=col)
            style = _cell_style(cell, theme_colors, scale)
            attrs = f' style="{style}"' if style else ""

            span = merge_map.get(key)
            if span:
                rs, cs = span
                if rs > 1:
                    attrs += f' rowspan="{rs}"'
                if cs > 1:
                    attrs += f' colspan="{cs}"'

            val = cell.value
            cell_ref = f"{get_column_letter(col)}{row}"
            if val is None:
                text = ""
            elif isinstance(val, str) and val.startswith("="):
                ext = extraction_map.get(cell_ref)
                if ext and ext.pattern != "unknown" and ext.placeholder:
                    text = ext.placeholder
                else:
                    # Unknown formula: render escaped text with a marker class
                    text = f'<span class="formula-unrecognized">{escape(val)}</span>'
            else:
                text = escape(str(val))

            img_key = (row - 1, col - 1)
            if img_key in image_map:
                text = image_map[img_key] + text

            lines.append(f"  <td{attrs}>{text}</td>")
        lines.append("</tr>")

    lines.append("</tbody>")
    lines.append("</table>")
    html = "\n".join(lines)
    if return_extractions:
        return html, [ext.to_dict() for ext in raw_extractions]
    return html


def csv_to_html(data: bytes) -> str:
    import pandas as pd
    df = pd.read_csv(BytesIO(data))
    return df.to_html(index=False, border=1)


def docx_to_html(data: bytes) -> str:
    import mammoth
    result = mammoth.convert_to_html(BytesIO(data))
    return result.value
