"""Extract and classify Excel formula patterns from template workbooks.

Parses known formula patterns (recommendation thresholds, simple recommendations,
unit conversions) into structured extractions. Unknown formulas are tagged as
"unknown" so the converter can flag them in the HTML output.
"""

import re
from dataclasses import dataclass, field

from openpyxl.utils import get_column_letter


@dataclass
class FormulaExtraction:
    cell_ref: str                       # e.g. "F34"
    formula: str                        # raw formula text
    pattern: str                        # "recommendation_threshold" | "simple_recommendation" | "unit_conversion" | "unknown"
    source_cell_ref: str | None = None  # the cell being tested (e.g. "E34")
    threshold: float | None = None
    operator: str | None = None         # "le", "lt", "ge", "gt"
    output_value: str | None = None     # "Needs Recs", etc.
    always_recommend: bool = False
    computation_params: dict | None = None  # for unit conversions
    placeholder: str | None = None      # suggested Jinja2 placeholder

    def to_dict(self) -> dict:
        return {
            "cell_ref": self.cell_ref,
            "formula": self.formula,
            "pattern": self.pattern,
            "source_cell_ref": self.source_cell_ref,
            "threshold": self.threshold,
            "operator": self.operator,
            "output_value": self.output_value,
            "always_recommend": self.always_recommend,
            "computation_params": self.computation_params,
            "placeholder": self.placeholder,
        }


# ---------------------------------------------------------------------------
# Cell ref helpers
# ---------------------------------------------------------------------------

def _cell_ref_to_key(ref: str) -> str:
    """Normalize a cell reference like '$E$34' or 'E34' to 'E34'."""
    return ref.replace("$", "").upper()


def _guess_analyte_key(source_ref: str, ws) -> str | None:
    """Try to derive an analyte key from the source cell's context.

    Looks at cells in the same row to find a label that could serve as the
    analyte name. Falls back to the column letter if nothing useful is found.
    """
    match = re.match(r"([A-Z]+)(\d+)", source_ref)
    if not match:
        return None
    col_letter = match.group(1)
    row = int(match.group(2))

    # Check columns A through D for a label in the same row
    for label_col in range(1, min(5, ord(col_letter) - ord("A") + 1)):
        cell = ws.cell(row=row, column=label_col)
        val = cell.value
        if isinstance(val, str) and val.strip() and not val.startswith("="):
            # Convert label to a safe key: "Phosphorus (P)" -> "P"
            # Try to extract parenthetical abbreviation first
            paren = re.search(r"\(([A-Za-z0-9_]+)\)", val)
            if paren:
                return paren.group(1)
            # Otherwise snake_case the label
            key = re.sub(r"[^a-zA-Z0-9]+", "_", val.strip()).strip("_")
            return key
    return col_letter


# ---------------------------------------------------------------------------
# Pattern matchers
# ---------------------------------------------------------------------------

def _try_recommendation_threshold(formula: str) -> dict | None:
    """Match: =IF(E34="","",IF((OR(E34<=11,E34="<1")), "Needs Recs","0"))

    Also handles simpler variants without OR, without the "<1" branch, or
    with different operators (<=, <, >=, >).
    """
    # Strip leading = and outer whitespace
    f = formula.lstrip("=").strip()

    # Pattern: IF(cell="","", IF( ... condition on cell vs threshold ... , "output", ...))
    # We look for a comparison of a cell ref against a numeric threshold
    # inside an IF that first checks for blank.

    # Find the cell reference being tested against blank
    blank_check = re.match(
        r'IF\(\s*(?:ISBLANK\(\s*([A-Z$]+\d+)\s*\)|([A-Z$]+\d+)\s*=\s*"")\s*,\s*""\s*,',
        f,
        re.IGNORECASE,
    )
    if not blank_check:
        return None

    source_ref = _cell_ref_to_key(blank_check.group(1) or blank_check.group(2))

    # Now look for the threshold comparison
    # Patterns: cell<=N, cell<N, cell>=N, cell>N (possibly inside OR())
    comparison = re.search(
        r'(?:' + re.escape(source_ref.replace("$", "")) + r'|\$?' + source_ref[0] + r'\$?' + source_ref[1:] + r')'
        r'\s*(<=|<|>=|>)\s*(\d+(?:\.\d+)?)',
        f,
        re.IGNORECASE,
    )
    if not comparison:
        # Try with dollar signs in the ref
        comparison = re.search(
            r'[A-Z$]+\d+\s*(<=|<|>=|>)\s*(\d+(?:\.\d+)?)',
            f,
            re.IGNORECASE,
        )
    if not comparison:
        return None

    op_map = {"<=": "le", "<": "lt", ">=": "ge", ">": "gt"}
    operator = op_map.get(comparison.group(1), "le")
    threshold = float(comparison.group(2))

    # Extract the output string (e.g. "Needs Recs")
    output_match = re.search(r',\s*"([^"]+)"\s*[,)]', f[blank_check.end():])
    output_value = output_match.group(1) if output_match else "Needs Recs"

    return {
        "pattern": "recommendation_threshold",
        "source_cell_ref": source_ref,
        "threshold": threshold,
        "operator": operator,
        "output_value": output_value,
    }


def _try_simple_recommendation(formula: str) -> dict | None:
    """Match: =IF(ISBLANK(E31),"","Needs Recs")

    Any formula that checks for blank and then always produces a fixed
    recommendation string (no threshold comparison).
    """
    f = formula.lstrip("=").strip()

    # ISBLANK variant
    m = re.match(
        r'IF\(\s*ISBLANK\(\s*([A-Z$]+\d+)\s*\)\s*,\s*""\s*,\s*"([^"]+)"\s*\)',
        f,
        re.IGNORECASE,
    )
    if m:
        return {
            "pattern": "simple_recommendation",
            "source_cell_ref": _cell_ref_to_key(m.group(1)),
            "output_value": m.group(2),
            "always_recommend": True,
        }

    # cell="" variant with no threshold
    m = re.match(
        r'IF\(\s*([A-Z$]+\d+)\s*=\s*""\s*,\s*""\s*,\s*"([^"]+)"\s*\)',
        f,
        re.IGNORECASE,
    )
    if m:
        return {
            "pattern": "simple_recommendation",
            "source_cell_ref": _cell_ref_to_key(m.group(1)),
            "output_value": m.group(2),
            "always_recommend": True,
        }

    return None


def _try_unit_conversion(formula: str) -> dict | None:
    """Match formulas that convert a cell value using multiplication/division.

    Example: =IF(E31="<1","0", E31*2*(C16/6)*(1000))/43560)

    These are computation formulas rather than simple lookups. We extract
    the source cell and flag the formula as a unit conversion so the admin
    can map it to a ComputedRecommendation entry.
    """
    f = formula.lstrip("=").strip()

    # Look for a pattern that multiplies/divides a cell reference
    # Must contain arithmetic operators beyond just a comparison
    has_arithmetic = bool(re.search(r'[A-Z$]+\d+\s*[*/]', f, re.IGNORECASE))
    if not has_arithmetic:
        return None

    # Extract the primary cell reference (first one that participates in math)
    ref_match = re.search(r'([A-Z$]+\d+)\s*[*/]', f, re.IGNORECASE)
    if not ref_match:
        return None

    source_ref = _cell_ref_to_key(ref_match.group(1))

    # Try to identify the conversion type from common patterns
    computation_type = "unit_conversion"
    params: dict = {"raw_formula": formula.lstrip("=")}

    # Nitrate-specific: presence of /43560 or *43560 suggests area conversion
    if "43560" in f:
        computation_type = "nitrate_conversion"
        params["description"] = "ppm to lb/1000 sq.ft. conversion"

    # Extract all cell references used in the formula
    all_refs = re.findall(r'[A-Z$]+\d+', f, re.IGNORECASE)
    params["referenced_cells"] = list(set(_cell_ref_to_key(r) for r in all_refs))

    return {
        "pattern": "unit_conversion",
        "source_cell_ref": source_ref,
        "computation_params": params,
        "computation_type": computation_type,
    }


# ---------------------------------------------------------------------------
# Placeholder generation
# ---------------------------------------------------------------------------

def _make_placeholder(pattern: str, source_ref: str, ws, output_value: str | None = None) -> str:
    """Generate a Jinja2 placeholder name from the extraction context."""
    analyte_key = _guess_analyte_key(source_ref, ws)
    if not analyte_key:
        analyte_key = source_ref

    # Sanitize for use in a Jinja2 variable name
    safe_key = re.sub(r"[^a-zA-Z0-9_]", "_", analyte_key).lower().strip("_")

    if pattern == "recommendation_threshold":
        return "{{ sample." + safe_key + "_recommendation }}"
    if pattern == "simple_recommendation":
        return "{{ sample." + safe_key + "_recommendation }}"
    if pattern == "unit_conversion":
        return "{{ sample." + safe_key + "_converted }}"
    return "{{ sample." + safe_key + "_formula }}"


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def extract_formulas(wb) -> list[FormulaExtraction]:
    """Scan all cells in the active worksheet for formulas and classify them.

    Returns a list of FormulaExtraction objects. Each formula cell gets exactly
    one entry. Known patterns get structured fields; unknown formulas get
    pattern="unknown" with just the raw text preserved.
    """
    ws = wb.active
    extractions: list[FormulaExtraction] = []

    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if not isinstance(val, str) or not val.startswith("="):
                continue

            cell_ref = f"{get_column_letter(cell.column)}{cell.row}"

            # Try each pattern matcher in priority order.
            # recommendation_threshold is checked before simple_recommendation
            # because the threshold pattern is a superset (it also checks blank).

            result = _try_recommendation_threshold(val)
            if result:
                ext = FormulaExtraction(
                    cell_ref=cell_ref,
                    formula=val,
                    pattern=result["pattern"],
                    source_cell_ref=result["source_cell_ref"],
                    threshold=result.get("threshold"),
                    operator=result.get("operator"),
                    output_value=result.get("output_value"),
                )
                ext.placeholder = _make_placeholder(
                    ext.pattern, ext.source_cell_ref, ws, ext.output_value
                )
                extractions.append(ext)
                continue

            result = _try_simple_recommendation(val)
            if result:
                ext = FormulaExtraction(
                    cell_ref=cell_ref,
                    formula=val,
                    pattern=result["pattern"],
                    source_cell_ref=result["source_cell_ref"],
                    output_value=result.get("output_value"),
                    always_recommend=result.get("always_recommend", False),
                )
                ext.placeholder = _make_placeholder(
                    ext.pattern, ext.source_cell_ref, ws, ext.output_value
                )
                extractions.append(ext)
                continue

            result = _try_unit_conversion(val)
            if result:
                ext = FormulaExtraction(
                    cell_ref=cell_ref,
                    formula=val,
                    pattern=result["pattern"],
                    source_cell_ref=result["source_cell_ref"],
                    computation_params=result.get("computation_params"),
                )
                ext.placeholder = _make_placeholder(
                    ext.pattern, ext.source_cell_ref, ws
                )
                extractions.append(ext)
                continue

            # Unknown formula
            extractions.append(FormulaExtraction(
                cell_ref=cell_ref,
                formula=val,
                pattern="unknown",
            ))

    return extractions
